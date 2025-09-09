#!/usr/bin/env python

# from ast import arg
from pathlib import Path
import pandas as pd
import numpy as np
import argparse
import re
import yaml


def get_args():
    parser = argparse.ArgumentParser(description=f"Crea la distribuzione dei posti in un'aula a partire dall'elenco degli studenti e dalla matrice dei posti disponibili.\n IMPORTANTE: le aule devono essere rappresentate con la cattedra in basso, le file numerate in maniera crescente, lasciando uno spazio nel caso in cui non sia presenta una fila.")
    parser.add_argument('-f', '--folder', type=Path, default=None, help="Percorso nel quale ricercare elenco studenti, file di configurazione e eventualmente excel con disposizioni, il nome del file finale avrà il nome della cartella come prefisso.")
    parser.add_argument('-r', '--rooms', type=str, default=None, help="Aule da considerare, separate da virgola. Indicare questo oppure inserire un file di configurazione YAML.")
    parser.add_argument('--nopc_students', type=str, default="", help="Lista delle matricole (senza spazi, separate da virgola) degli studenti che necessitano di essere posizionati in aula multimediale.")
    parser.add_argument('--nopc_room', type=str, default="5T", help="Aula multimediale nella quale inserire nopc_students. Default è '5T'.")
    parser.add_argument('--dsa_room', type=str, default=None, help="Aula nella quale saranno posizionati gli studenti DSA.")
    parser.add_argument('-n', '--name', type=str, default="COD_yyyymmdd", help="Prefisso del nome dei files di output. Default: '<COD>_yyyymmdd'.")
    parser.add_argument("--random_order", action="store_true", help="Randomize order of students instead of alphabetical.")
    parser.add_argument("--nostyle", action="store_true", help="Prevent adding style to the resulting sheet.")
    args = parser.parse_args()
    return args


def snake_j(row, i, j):
    rev_j = row.index[len(row) - row.index.get_loc(j) - 1]
    if i % 4 == 2 and not np.isnan(row[rev_j]):
         return rev_j
    return j


def stamp_id(room_name, config, matricole, prenotati):
    window = config["position"]

    # B2:U20 is the format of config["position"] (the letter can also be AA, AB, etc.)
    
    rows = list(map(int, re.findall(r"\d+", window)))
    cols = re.findall(r"[A-Z]+", window)
    cols = [ord(c[-1].lower())-97 + 26*(len(c)-1) for c in cols]

    if "filename" in config:
        if config["filename"].name.endswith(".csv"):
            aula = pd.read_csv(config["filename"], sep="\t", header=None)
        elif config["filename"].name.endswith(".xlsx"):
            aula = pd.read_excel(config["filename"], sheet_name=config["sheet"], header=None)
        
    else: # retrieve from Google's API
        # this might return a wrong CSV, beware
        sheet_id = "1yuDN40n8pcoP2FgjUwuR6oxKvNRSa2jbJTVhOrsB5_A"
        sheet_name = config["sheet"].replace(" ", "%20")
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"
        aula = pd.read_csv(
            url,
            header=None
        )
        
    col_names = aula.loc[rows[0]-1, cols[0]+1: cols[1]]  
    aula = aula.iloc[rows[0]: rows[1], cols[0]: cols[1] + 1].copy()
    aula = aula.set_index(aula.columns[0]).rename_axis(None)
    aula.columns = col_names

    aula = aula.map(lambda x: np.nan if x==0 else x)
    slots = (~aula.isna()).sum().sum()
        
    print(f"| Room {room_name.ljust(3)} | Slots: {slots}, remaining students: {len(matricole)}")

    if len(matricole) < np.nansum(aula.values):
        matricole += ["x" for _ in range(int(np.nansum(aula.values)) - len(matricole))]
    
        # Allow one column to have an empty name (first unnamed column)
        unnamed_cols = [c for c in aula.columns if type(c)==str and "Unnamed" in c]
        if unnamed_cols:
            aula = aula.rename(columns={unnamed_cols[0]: ""})
            # Optionally, keep other unnamed columns as is or handle as needed
    aula = aula.rename(index = {i: chr(int(i)+64) if i>0 else np.nan for i in aula.index})
        
    placement = aula.copy()
    
    for i, row in aula.sort_index().iterrows():
    
        row = row.astype(float)
        if not matricole:
            break
        
        for j, place in row.items():
            if not matricole:
                break
                
            if place == 1: # force notation here
                curr_matricola = matricole.pop(0) 
                sn_j = snake_j(row, ord(i), j) 
                placement.loc[i, sn_j] = float(curr_matricola) if not curr_matricola == "x" else -1.
                
                if not curr_matricola == "x":
                    prenotati.loc[curr_matricola, ["AULA", "POSTO"]] = str(room_name), f"{i}{int(sn_j)}"

    placement = placement.fillna("")
    placement[""] = ""
    
    # Center the professor desk, covering multiple columns
    n_cols = placement.shape[1] - 1
    desk_row = {}
    if n_cols % 2 == 1:
        # Odd: cover central, one before, one after
        center = n_cols // 2
        for idx, c in enumerate(placement.columns):
            if idx in [center-1, center, center+1]:
                desk_row[c] = "Professor Desk"
            else:
                desk_row[c] = ""
    else:
        # Even: cover the two central columns
        center1 = n_cols // 2 - 1
        center2 = n_cols // 2
        for idx, c in enumerate(placement.columns):
            if idx in [center1, center2]:
                desk_row[c] = "Professor Desk"
            else:
                desk_row[c] = ""
    # add one empty row before the desk
    placement.loc["   "] = ""
    placement.loc["      "] = pd.Series(desk_row)
    return placement, matricole


def styled_seats(x):
    checkerboard = lambda d: np.vstack(d[0]*(np.r_[d[1]*[True,False]], np.r_[d[1]*[False,True]]))[:d[0], :d[1]]
    def style_cell(val, row_idx, col_idx):
        col = x.columns[col_idx]
        if isinstance(col, str) and col.strip() == "":
            return 'width: 50px; text-align: center'
        if pd.notna(val) and checkerboard(x.shape)[row_idx, col_idx]:
            return 'background-color: lightgrey; width: 50px; text-align: center'
        return 'width: 50px; text-align: center'
    df1 = pd.DataFrame(
        [[style_cell(x.iloc[i, j], i, j) for j in range(x.shape[1])] for i in range(x.shape[0])],
        index=x.index, columns=x.columns
    )
    # blank style to all the rows with empty string as index
    df1.loc[x.index.str.strip() == ""] = ['' for _ in range(x.shape[1])]
    df1.loc[x.index.isna()] = ['' for _ in range(x.shape[1])]
    df1.iloc[-1] = ["" if x == "" else 'background-color: lightgrey; text-align: center' for x in x.iloc[-1]]
    return df1


def main(args):
    args.yconfig = args.folder / "riferimenti_aule.yaml"
    args.name = args.folder / args.folder.stem        

    if args.yconfig.exists():
        with open(args.yconfig, "r") as f:
            riferimenti = yaml.load(f, yaml.SafeLoader)
            riferimenti = {str(k): v for k, v in riferimenti.items()}
    elif args.rooms is not None:
        with open("riferimenti_aule_all.yaml", "r") as f:
            aule_db = yaml.load(f, yaml.SafeLoader)
        def process_room(room):
            try:
                return int(room)
            except:
                return room
            
        riferimenti = {a: aule_db.get(a, aule_db.get(process_room(a))) for a in args.rooms.split(",")}
    else:
        raise FileNotFoundError(f"File {args.yconfig} not found. Please create a conf file or provide a list of rooms.") 

    prenotati = pd.DataFrame()
    for p in args.folder.glob("VISAP_Elenco_Studenti_*"):
        if p.suffix == ".xlsx" or p.suffix == ".xls": 
            tmp = pd.read_excel(p) 
            row_toskip = (tmp.iloc[:,0]=="#").argmax()
            col = tmp.xs(row_toskip)
            tmp = tmp.drop(range(row_toskip+1)).reset_index(drop=True)
            tmp.columns = col
        else:
            tmp = pd.read_csv(p)
        prenotati = pd.concat([prenotati,tmp], ignore_index=True)
        
    plen = len(prenotati)
    prenotati = prenotati.drop_duplicates()
    if plen != len(prenotati): print("Attenzione: studenti duplicati nella tabella\n")

    prenotati.columns = [c.strip() for c in prenotati.columns]
    prenotati = prenotati.drop(["DATA PRENOTAZIONE", "DOMANDA", "RISPOSTA", "CORSO", "CDL", "NUMERO CORSO"], axis=1)
    
    if args.random_order:
        prenotati = prenotati.sample(frac=1).set_index("MATRICOLA")
    else:
        prenotati = prenotati.sort_values(by="COGNOME").set_index("MATRICOLA")
    
    prenotati = prenotati.assign(AULA=np.nan, POSTO=np.nan)
    prenotati.AULA = prenotati.AULA.astype('object')
    prenotati.POSTO = prenotati.POSTO.astype('object')
    
    # prenotati.NOTE = prenotati.NOTE.astype('str')
    if not prenotati.NOTE.dtype == 'object': prenotati.NOTE = ""
    matricole = prenotati[~prenotati.NOTE.str.contains("Esame online", na=False)].index.to_list()

    if not args.dsa_room:
        args.dsa_room = list(riferimenti.keys())[0]
    matricole_dsa = prenotati[prenotati.NOTE.str.contains("Dsa", na=False)|prenotati.NOTE.str.contains("Tempo aggiuntivo", na=False)].index.to_list()
    print(f"\nDsa students ({', '.join(prenotati[prenotati.index.isin(matricole_dsa)].COGNOME)}) will be placed in room {args.dsa_room}\n")
    matricole = [m for m in matricole if m not in matricole_dsa]
    
    if "nopc_students" in riferimenti.keys():
        args.nopc_students = riferimenti.pop("nopc_students")
    if args.nopc_students:
        ex_students = [int(m) for m in str(args.nopc_students).split(",")]
        prenotati.loc[ex_students, "AULA"] = args.nopc_room
        matricole = [m for m in matricole if m not in ex_students]
        print(f"These students will be placed in {args.nopc_room}:")
        print(ex_students, "\n")
    else:
        ex_students = []

    writer_d = pd.ExcelWriter(f"{args.name}_disposizioni.xlsx")
    writer_p = pd.ExcelWriter(f"{args.name}_prenotati.xlsx")
    
    
    args.rooms = list(riferimenti.keys())
        
    if len(args.rooms) > 1 and len(set(args.rooms)) == 1:
        room_names = [f"{i+1}{r}" for i, r in enumerate(args.rooms)]
    else:
        room_names = args.rooms
        
    print("Room names:", room_names, "\n")
    
    for i, (a, room_name) in enumerate(zip(args.rooms, room_names)):
        config = riferimenti.get(a)
        if "position" not in config:
            raise RuntimeError("Specify a position window in YAML reference file!")
        if "filename" not in config and args.folder:
            config["filename"] = args.folder / "Aule esame.xlsx"
            if not config["filename"].exists(): config["filename"] = Path.cwd() / "Aule esami Polito, disposizione posti.xlsx"
        
        if args.dsa_room is not None and room_name==args.dsa_room:
            print(f"Placing dsa students in {room_name}")
            matricole = matricole_dsa + matricole
        aula, matricole = stamp_id(room_name, config, matricole, prenotati)
        
        aula = aula.map(lambda x: 'x' if x==-1 else x).map(lambda x: int(x) if isinstance(x, float) else x)
        if not args.nostyle:
            aula = aula.rename(columns={c: j * " " if pd.isnull(c) and c != "" else c for j, c in enumerate(aula.columns)})
            aula = aula.style.apply(styled_seats, axis=None)
        
        aula.to_excel(writer_d, sheet_name=f"Aula_{room_name}")

        prenotati[prenotati.AULA == room_name].to_excel(writer_p, sheet_name=f"Aula_{room_name}")

    # remove placeholders "x"
    # matricole = [m for m in matricole if m != "x"]
    if len(matricole) != 0:
        if len(matricole) + len(ex_students) < 1: # Se pochi, vengono caricati in 5T: NON PIU' DISPONIBILE
            prenotati.loc[matricole, "AULA"] = args.nopc_room
            print(f"\nℹ️  {len(matricole)} extra students are placed in {args.nopc_room}")
        else:
            raise ValueError(f"⚠️  Designated rooms are not sufficient, {len(matricole)} students missing ({matricole})")

    if args.nopc_students:
        prenotati[prenotati.AULA == args.nopc_room].to_excel(writer_p, sheet_name=f"Aula_{args.nopc_room}")
        
    if len(room_names) > 1 or args.nopc_students:
        prenotati.to_excel(writer_p, sheet_name=f"Elenco Complessivo")
        
    limiti = prenotati[prenotati.AULA != args.nopc_room].groupby("AULA").agg({
            "COGNOME": ["min", "max"]
        })
    print("\n✔️  Students succesfully allocated\n", str(limiti))

    writer_d.close()
    writer_p.close()

    # Merge 'Professor Desk' cells in the exported Excel file (after closing writer)
    import openpyxl
    wb = openpyxl.load_workbook(f"{args.name}_disposizioni.xlsx")
    for room_name in (room_names):
        if f"Aula_{room_name}" in wb.sheetnames:
            ws = wb[f"Aula_{room_name}"]
            for row in ws.iter_rows():
                desk_cells = [cell for cell in row if cell.value == "Professor Desk"]
                if desk_cells:
                    indices = [cell.column for cell in desk_cells]
                    if indices:
                        start = min(indices)
                        end = max(indices)
                        ws.merge_cells(start_row=desk_cells[0].row, start_column=start,
                                       end_row=desk_cells[0].row, end_column=end)
                    break
    wb.save(f"{args.name}_disposizioni.xlsx")
        
    
if __name__ == '__main__':
    args = get_args()
    main(args)



    
